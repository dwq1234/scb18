
#接口自动化测试
'''
1.excel测试用例准备ok，代码可以自动读取用例数据
2.执行接口测试，得到响应结果
3，断言：响应结果==预期结果    判断通过还是不通过
4，写入入最终执行通过与否的结果----excel表格
'''

import requests
import openpyxl
#把接口请求封装成函数
def api_fun(url,data):    #定义会变化得参数，地址和请求体
    header_reg = {"X-Lemonban-Media-Type":"lemonban.v2","Content-Type":"application/json"} #请求头
    res_reg = requests.post(url=url,json=data,headers=header_reg).json()  #可以使用ctrl点击进行查看相关用法,定义一个变量resultl来接收函数的返回值
    return res_reg   #定义一个返回值
url_reg = 'http://120.78.128.25:8766/futureloan/member/register'   #定义一个变量为url
data_reg = {"mobile_phone": "15922147763","pwd": "lemon1234", "type":"1","reg_name":"lemon"}  #定义一个变量为请求体
# res_reg = api_fun(url = url_reg,data = data_reg)   #调用函数
# print(res_reg)

#读取定义为函数，进行调用
def read_case(filename,sheetname):   #定义函数，将会变的值设定为参数
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheetname]
    rows_max = sheet.max_row   #获取最大行数
    # print(rows_max)
    data_list = []   #定义一个空列表,存放for循环依次读取到得测试用例数据
    for i in range(2, rows_max+1):  #获取最大行数,取头不取尾需要+1
        data_dict = dict( case_id =sheet.cell(row=i,column=1).value,
        url = sheet.cell(row=i,column=5).value,  #
        data = sheet.cell(row=i,column=6).value,
        excepted = sheet.cell(row=i,column=7).value, )
        data_list.append(data_dict)   #将遍历得到得数据追加在空列表里面
    return data_list
# read_result = read_case('test_case_api.xlsx','login')  #定义一个变量接收返回值
# print(read_result)  #打印最终结果
#写入函数
def write_file(filename,sheetname,row,column,fial_result):
    workbook = openpyxl.load_workbook(filename)
    sheet = workbook[sheetname]
    sheet.cell(row=row,column=column).value = fial_result
    workbook.save(filename)
# write_file("test_case_api.xlsx",'register',6,8,'Passed')   #函数调'''用

#eval函数
'''
1.运行被字符串包裹着得表达式
2.eval('{"mobile_phone":"13652440101","pwd":"12345678","type":1,"reg_name":"34254sdfs"}'),运行之后，字符串就会变成字典格式
dict1 = (eval('{"mobile_phone":"13652440101","pwd":"12345678","type":1,"reg_name":"34254sdfs"}'))#,运行之后，字符串就会变成字典格式
print(dict1)
print(type(dict1))
3.如果里面是一个表达式，会自动计算然后得出结果
eg:
print(eval('2+9'))
'''
# 1.调用读取测试数据函数
'''test_read = read_case("../test_data/test_case_api.xlsx", 'login')
# print(test_read)   #得到一个列表，里面有13条字典类型的测试数据
#进行用例执行的时候，一条一条的读取里面的测试数据，然后把列表里的地址，请求数据以及期望值取出来。
for num in test_read:    #遍历列表数据,即把字典格式得测试数据从列表里取出来
    print(num)
    test_id = num['case_id']
    test_excepted = num['excepted']    #获取期望信息，后面用于和实际结果进行比对
    # print(test_excepted)
    test_url = num['url']  #读取到的数据是字典类型，将字典里的地址取出来
    test_data =eval(num['data'])  #取出data,但是取出出来的数据类型是字符串，，data必须传字典格式，所以得用eval()函数把字符串转换为字典
    # print(test_data)
    # print(type(test_data))
    test_excepted =eval(num['excepted'])
    # print(test_excepted)
#2已经将数据取出来，开始进行用例执行，调用用例执行函数
    test_result = api_fun(test_url,test_data)
    print(test_result)
#执行了测试用例之后需要将预期结果与实际结果进行比较以此来判定测试是否通过，所以要先把返回得实际结果取出来，返回结果为字典格式，用dict[]进行取即可
    excepted_msg = test_excepted['msg']   #获取期望的msg信息
    excepted_code = test_excepted['code']  #获取期望的code信息
    result_msg = test_result['msg']   #获取实际的msg信息
    result_code = test_result['code']   #获取实际的code信息
    print(test_id)
    print('期望code为:{}'.format(excepted_code),'msg为:{}'.format( excepted_msg))
    # print('msg为:{}'.format( excepted_msg))
    print('实际code为:{}'.format(result_code),'msg为:{}'.format(result_msg))
    # print('msg为:{}'.format(result_msg))
#把实际code和msg与实际code和msg取出来之后，就可以进行判断,如果相等，测试用例表格写入，passed,如果失败，写入faile
    if result_code  == excepted_code and result_msg == excepted_msg:
        print('第{}条测试用例通过'.format(test_id))
        write_file('../test_data/test_case_api.xlsx', 'login', test_id + 1, 8, 'Passed')    #调用写入结果函数，将测试结果写入文件
    else:
        print('第{}条测试用例不通过'.format(test_id))
        write_file('../test_data/test_case_api.xlsx', 'login', test_id + 1, 8, 'Falile')   #调用写入结果函数，将测试结果写入文件'''


#将整个结果封装为一个函数，可以方便进行调用
def test_fun(filename,sheetname):
    test_read = read_case(filename, sheetname)
    # print(test_read)   #得到一个列表，里面有13条字典类型的测试数据
    # 进行用例执行的时候，一条一条的读取里面的测试数据，然后把列表里的地址，请求数据以及期望值取出来。
    for num in test_read:  # 遍历列表数据,即把字典格式得测试数据从列表里取出来
        # print(num)
        test_id = num['case_id']
        test_excepted = num['excepted']  # 获取期望信息，后面用于和实际结果进行比对
        # print(test_excepted)
        test_url = num['url']  # 读取到的数据是字典类型，将字典里的地址取出来
        test_data = eval(num['data'])  # 取出data,但是取出出来的数据类型是字符串，，data必须传字典格式，所以得用eval()函数把字符串转换为字典
        # print(test_data)
        # print(type(test_data))
        test_excepted = eval(num['excepted'])
        # print(test_excepted)
        # 2已经将数据取出来，开始进行用例执行，调用用例执行函数
        test_result = api_fun(test_url, test_data)
        # print(test_result)
        # 执行了测试用例之后需要将预期结果与实际结果进行比较以此来判定测试是否通过，所以要先把返回得实际结果取出来，返回结果为字典格式，用dict[]进行取即可
        excepted_msg = test_excepted['msg']  # 获取期望的msg信息
        excepted_code = test_excepted['code']  # 获取期望的code信息
        result_msg = test_result['msg']  # 获取实际的msg信息
        result_code = test_result['code']  # 获取实际的code信息
        print(test_id)
        print('期望code为:{}'.format(excepted_code), 'msg为:{}'.format(excepted_msg))
        # print('msg为:{}'.format( excepted_msg))
        print('实际code为:{}'.format(result_code), 'msg为:{}'.format(result_msg))
        # print('msg为:{}'.format(result_msg))
        # 把实际code和msg与实际code和msg取出来之后，就可以进行判断,如果相等，测试用例表格写入，passed,如果失败，写入faile
        if result_code == excepted_code and result_msg == excepted_msg:
            print('第{}条测试用例通过'.format(test_id))
            write_file(filename, sheetname, test_id + 1, 8, 'Passed')  # 调用写入结果函数，将测试结果写入文件
        else:
            print('第{}条测试用例不通过'.format(test_id))
            write_file(filename, sheetname, test_id + 1, 8, 'Falile')  # 调用写入结果函数，将测试结果写入文件
#调用函数,定义一个变量接收函数返回值
test_result = test_fun("C:\\python_project\\python\\python\\test_data\\test_case_api.xlsx",'login')
test_result = test_fun("C:\\python_project\\python\\python\\test_data\\test_case_api.xlsx",'register')
print(test_result)








